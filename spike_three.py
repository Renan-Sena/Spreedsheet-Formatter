import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def processar_aba(df, ws):
    # Trata a coluna "Data"
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors='coerce')
        if df["Data"].isnull().any():
            print("⚠️ Algumas datas são inválidas e foram convertidas para vazio (NaT).")

    # Adiciona os dados no worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Linha onde os totais serão adicionados
    total_row = ws.max_row + 1
    ws[f"A{total_row}"] = "Totais"
    ws[f"B{total_row}"] = f"=SUBTOTAL(109,B2:B{total_row - 1})"
    ws[f"C{total_row}"] = f"=SUBTOTAL(109,C2:C{total_row - 1})"
    ws[f"D{total_row}"] = ""

    # Cabeçalho em negrito
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Adiciona tabela com estilo e filtros
    table_ref = f"A1:D{total_row}"
    tab = Table(displayName=f"Tabela_{ws.title}", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2


def gerar_planilha(arquivo_entrada, arquivo_saida="Planilha_Custos_Empresas.xlsx"):
    try:
        print(f"🔍 Lendo planilhas do arquivo: {arquivo_entrada}")
        planilhas = pd.read_excel(arquivo_entrada, sheet_name=None)
        print(f"✅ {len(planilhas)} aba(s) encontrada(s): {list(planilhas.keys())}")
    except Exception as e:
        print(f"❌ Erro ao ler o arquivo de entrada: {e}")
        return

    wb = Workbook()
    wb.remove(wb.active)  # Remove a aba padrão vazia

    for nome_aba, df in planilhas.items():
        print(f"➡️ Processando aba: {nome_aba}")
        # Verifica se colunas mínimas existem
        colunas_esperadas = {"Data", "Empresa 1 (R$)", "Empresa 2 (R$)"}
        if not colunas_esperadas.issubset(df.columns):
            print(f"⚠️ Ignorando aba '{nome_aba}' – colunas esperadas não encontradas.")
            continue

        ws = wb.create_sheet(title=nome_aba[:31])  # Limita nome da aba a 31 caracteres
        processar_aba(df, ws)

    if not wb.sheetnames:
        print("❌ Nenhuma aba válida foi processada. Nenhum arquivo gerado.")
        return

    try:
        wb.save(arquivo_saida)
        print(f"✅ Planilha gerada com sucesso: {arquivo_saida}")
    except Exception as e:
        print(f"❌ Erro ao salvar a planilha: {e}")


# Execução via terminal
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("⚠️ Uso: python seu_script.py entrada.xlsx [saida.xlsx]")
    else:
        entrada = sys.argv[1]
        saida = sys.argv[2] if len(sys.argv) > 2 else "Planilha_Custos_Empresas.xlsx"
        gerar_planilha(entrada, saida)

# Exemplo de execução: python seu_script.py entrada.xlsx(nome de entrada) relatorio_final.xlsx(nome de sáida)
