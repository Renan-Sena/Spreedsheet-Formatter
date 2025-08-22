import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from PySide6 import QtCore
from PySide6.QtWidgets import (
    QMainWindow, 
    QWidget, 
    QVBoxLayout, 
    QHBoxLayout,
    QPushButton, 
    QFileDialog, 
    QMessageBox, 
    QTableView, 
    QStatusBar,
    QDialog, 
    QHeaderView
)
from PySide6.QtGui import QIcon

from models.pandas_model import PandasModel
from dialogs.insert_rows_dialog import InsertRowsDialog
from dialogs.remove_rows_dialog import RemoveRowsDialog
from dialogs.plain_text_dialog import PlainTextDialog
from dialogs.filter_dialog import FilterDialog


class SpreadSheetApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Spreadsheet Editor (PySide6)")
        self.resize(1100, 700)

        self.df = pd.DataFrame()
        self.sort_direction = {}
        self.file_loaded = False

        central = QWidget()
        self.setCentralWidget(central)
        vbox = QVBoxLayout(central)
        vbox.setContentsMargins(12, 12, 12, 12)

        hb = QHBoxLayout()
        hb.setSpacing(12)

        def resource_path(relative_path):
            """
            Returns the correct path for resources, works in terminal and PyInstaller exe
            """
            try:
                base_path = sys._MEIPASS  # PyInstaller executable
            except AttributeError:
                base_path = os.path.abspath(".")  # terminal
            return os.path.join(base_path, relative_path)

        # Icons directory
        ICONS_DIR = resource_path("public/icons")  # just the folder, don't include individual files

        def make_btn(text: str, icon_file: str, slot):
            icon_path = os.path.join(ICONS_DIR, icon_file)
            btn = QPushButton(text)

            if os.path.exists(icon_path):
                btn.setIcon(QIcon(icon_path))
                btn.setIconSize(QtCore.QSize(20, 20))
            else:
                btn.setToolTip(f"Icon not found: {icon_path}")

            btn.clicked.connect(slot)
            btn.setMinimumWidth(160)
            return btn

        hb.addWidget(make_btn("Open Spreadsheet", "folder-open-solid-full.svg", self.open_file))
        hb.addWidget(make_btn("Insert Row", "grip-lines-solid-full.svg", self.add_row))
        hb.addWidget(make_btn("Remove Row", "delete-left-solid-full.svg", self.remove_row))
        hb.addWidget(make_btn("Add Column", "table-columns-solid-full.svg", self.add_column))
        hb.addWidget(make_btn("Remove Column", "trash-can-solid-full.svg", self.remove_column))
        hb.addWidget(make_btn("Filters", "filter-solid-full.svg", self.edit_column_filters))
        hb.addStretch()
        hb.addWidget(make_btn("Save", "floppy-disk-solid-full.svg", self.save_spreadsheet))
        vbox.addLayout(hb)

        self.view = QTableView()
        self.view.setAlternatingRowColors(True)
        self.view.horizontalHeader().setStretchLastSection(True)
        self.view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.view.setSelectionBehavior(QTableView.SelectRows)
        self.view.setEditTriggers(QTableView.DoubleClicked | QTableView.EditKeyPressed | QTableView.AnyKeyPressed)
        self.view.horizontalHeader().sectionClicked.connect(self._sort_by_header)
        vbox.addWidget(self.view)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self._update_status()

        self.model = PandasModel(self.df, self)
        self.view.setModel(self.model)

    def _update_status(self):
        if self.df.empty:
            self.status.showMessage("No spreadsheet loaded")
        else:
            self.status.showMessage(f"Rows: {len(self.df)}    Columns: {len(self.df.columns)}")

    def _reload_model(self):
        self.model.update_df(self.df)
        self._update_status()
        self.view.resizeColumnsToContents()

    # --------- Actions ----------
    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel file",
                                              "", "Excel Spreadsheets (*.xls *.xlsx)")
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            engine = "openpyxl" if ext == ".xlsx" else "xlrd"
            self.df = pd.read_excel(path, engine=engine)
            self.df.attrs.setdefault("filters", {})
            self.file_loaded = True
            self.sort_direction.clear()
            self._reload_model()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error reading file:\n{e}")

    def _sort_by_header(self, logical_index: int):
        if self.df.empty or logical_index < 0 or logical_index >= len(self.df.columns):
            return
        column = self.df.columns[logical_index]
        direction = self.sort_direction.get(column, False)
        try:
            self.df = self.df.sort_values(by=column, ascending=not direction).reset_index(drop=True)
            self.sort_direction[column] = not direction
            self._reload_model()
        except Exception as e:
            QMessageBox.warning(self, "Warning", f"Could not sort by '{column}':\n{e}")

    def add_row(self):
        if self.df.empty:
            QMessageBox.warning(self, "Warning", "Open a spreadsheet first.")
            return

        dlg = InsertRowsDialog(self.df.columns, self)
        if dlg.exec() != QDialog.Accepted:
            return

        pos = dlg.pos
        if pos < 0 or pos >= len(self.df):
            QMessageBox.warning(self, "Warning", "Index out of range.")
            return

        new_row = pd.DataFrame([dlg.values], columns=self.df.columns)

        if dlg.before:
            part1 = self.df.iloc[:pos]
            part2 = self.df.iloc[pos:]
        else:
            part1 = self.df.iloc[:pos+1]
            part2 = self.df.iloc[pos+1:]

        self.df = pd.concat([part1, new_row, part2], ignore_index=True)
        self._reload_model()

    def remove_row(self):
        if self.df.empty:
            QMessageBox.warning(self, "Warning", "Open a spreadsheet first.")
            return

        dlg = RemoveRowsDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return

        idx = dlg.idx
        if idx < 0 or idx >= len(self.df):
            QMessageBox.warning(self, "Warning", "Index out of range.")
            return

        self.df = self.df.drop(index=idx).reset_index(drop=True)
        self._reload_model()

    def add_column(self):
        if self.df.empty:
            QMessageBox.warning(self, "Warning", "Open a spreadsheet first.")
            return

        dlg = PlainTextDialog("Add Column", "Name of new column:", self)
        if dlg.exec() != QDialog.Accepted:
            return
        name = dlg.value
        if name in self.df.columns:
            QMessageBox.warning(self, "Warning", "Column already exists.")
            return
        self.df[name] = ""
        self._reload_model()

    def remove_column(self):
        if self.df.empty:
            QMessageBox.warning(self, "Warning", "Open a spreadsheet first.")
            return

        dlg = PlainTextDialog("Remove Column", "Name of column to remove:", self)
        if dlg.exec() != QDialog.Accepted:
            return
        name = dlg.value
        if name not in self.df.columns:
            QMessageBox.warning(self, "Warning", "Column does not exist.")
            return
        self.df = self.df.drop(columns=[name])
        self._reload_model()

    def edit_column_filters(self):
        if self.df.empty:
            QMessageBox.warning(self, "Warning", "Open a spreadsheet first.")
            return

        current_filters = self.df.attrs.get("filters", {})
        dlg = FilterDialog(self.df.columns, current_filters, self)
        new_filters = dlg.get_filters()
        if new_filters is None:
            return

        try:
            filtered_df = self.df.copy()
            for col, val in new_filters.items():
                filtered_df = filtered_df[filtered_df[col].astype(str).str.contains(val, case=False, na=False)]
            self.df.attrs["filters"] = new_filters
            self.df = filtered_df.reset_index(drop=True)
            self._reload_model()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error applying filters:\n{e}")

    def save_spreadsheet(self):
        if self.df.empty:
            QMessageBox.warning(self, "Warning", "No spreadsheet to save.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Save spreadsheet", "", "Excel Spreadsheets (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Spreadsheet"

            # Bold headers
            for idx, col in enumerate(self.df.columns, 1):
                cell = ws.cell(row=1, column=idx, value=col)
                cell.font = Font(bold=True)

            # Data
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Styled table
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tab = Table(displayName="Table1", ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            wb.save(path)
            QMessageBox.information(self, "Success", "Spreadsheet saved successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving file:\n{e}")
