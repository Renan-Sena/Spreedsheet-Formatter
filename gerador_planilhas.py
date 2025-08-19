import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from PySide6 import QtCore
from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QMessageBox, QTableView, QStatusBar,
    QDialog, QLabel, QLineEdit, QFormLayout, QDialogButtonBox,
    QRadioButton, QGridLayout, QScrollArea, QHeaderView
)
from PySide6.QtGui import QAction, QIcon


class PandasModel(QAbstractTableModel):
    def __init__(self, df: pd.DataFrame, parent=None):
        super().__init__(parent)
        self._df = df

    @property
    def df(self):
        return self._df

    def update_df(self, df: pd.DataFrame):
        self.beginResetModel()
        self._df = df
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self._df)

    def columnCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self._df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        r, c = index.row(), index.column()
        value = self._df.iat[r, c]

        if role in (Qt.DisplayRole, Qt.EditRole):
            if pd.api.types.is_datetime64_any_dtype(self._df.dtypes[c]):
                if pd.isna(value):
                    return ""
                try:
                    return pd.to_datetime(value).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    return str(value) if value is not None else ""
            return "" if pd.isna(value) else str(value)
        return None

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable

    def setData(self, index, value, role=Qt.EditRole):
        if role != Qt.EditRole or not index.isValid():
            return False
        r, c = index.row(), index.column()
        col = self._df.columns[c]
        dtype = self._df[col].dtype

        try:
            if pd.api.types.is_integer_dtype(dtype):
                self._df.at[r, col] = int(value) if value != "" else pd.NA
            elif pd.api.types.is_float_dtype(dtype):
                self._df.at[r, col] = float(value) if value != "" else pd.NA
            elif pd.api.types.is_datetime64_any_dtype(dtype):
                self._df.at[r, col] = pd.to_datetime(value) if value != "" else pd.NaT
            else:
                self._df.at[r, col] = value
        except Exception:
            return False

        self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
        return True

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return str(self._df.columns[section]) if section < len(self._df.columns) else ""
        else:
            return str(section)


class FiltrosDialog(QDialog):
    def __init__(self, columns, filtros_existentes=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Editar Filtros")
        self.setMinimumSize(420, 480)

        layout = QVBoxLayout(self)

        info = QLabel("Defina os filtros para as colunas (vazio = sem filtro).")
        info.setWordWrap(True)
        layout.addWidget(info)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        form = QFormLayout(inner)
        self.inputs = {}

        for col in columns:
            le = QLineEdit()
            if filtros_existentes and col in filtros_existentes:
                le.setText(str(filtros_existentes[col]))
            self.inputs[col] = le
            form.addRow(QLabel(col + ":"), le)

        scroll.setWidget(inner)
        layout.addWidget(scroll)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_filters(self):
        """Abre o diálogo e retorna um dict {coluna: texto} apenas com valores não vazios.
        Retorna None se o usuário cancelar."""
        if self.exec() != QDialog.Accepted:
            return None
        result = {}
        for col, le in self.inputs.items():
            txt = le.text().strip()
            if txt:
                result[col] = txt
        return result


class InserirLinhaDialog(QDialog):
    def __init__(self, columns, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Adicionar Linha")
        self.setMinimumSize(420, 520)
        self.valores = None
        self.pos = None
        self.antes = False

        layout = QVBoxLayout(self)

        gridw = QWidget()
        grid = QGridLayout(gridw)
        self.inputs = {}
        for i, col in enumerate(columns):
            grid.addWidget(QLabel(col), i, 0)
            le = QLineEdit()
            self.inputs[col] = le
            grid.addWidget(le, i, 1)

        layout.addWidget(gridw)

        pos_row = QHBoxLayout()
        pos_row.addWidget(QLabel("Inserir linha antes/depois de qual índice?"))
        self.pos_edit = QLineEdit()
        self.pos_edit.setPlaceholderText("Ex.: 0, 1, 2 ...")
        pos_row.addWidget(self.pos_edit)
        layout.addLayout(pos_row)

        rb_row = QHBoxLayout()
        self.rb_antes = QRadioButton("Antes")
        self.rb_depois = QRadioButton("Depois")
        self.rb_depois.setChecked(True)
        rb_row.addWidget(self.rb_antes)
        rb_row.addWidget(self.rb_depois)
        layout.addLayout(rb_row)

        self.err = QLabel("")
        self.err.setStyleSheet("color:#ff5555; font-weight:bold;")
        layout.addWidget(self.err)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._ok)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _ok(self):
        pos_txt = self.pos_edit.text().strip()
        if not pos_txt.isdigit():
            self.err.setText("Índice precisa ser um número inteiro válido.")
            return
        self.pos = int(pos_txt)
        self.antes = self.rb_antes.isChecked()

        vals = {col: le.text() for col, le in self.inputs.items()}
        self.valores = vals
        self.accept()


class RemoverLinhaDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Remover Linha")
        self.idx = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Informe o índice da linha a remover:"))
        self.le = QLineEdit()
        layout.addWidget(self.le)
        self.err = QLabel("")
        self.err.setStyleSheet("color:#ff5555; font-weight:bold;")
        layout.addWidget(self.err)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._ok)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _ok(self):
        txt = self.le.text().strip()
        if not txt.isdigit():
            self.err.setText("Índice inválido.")
            return
        self.idx = int(txt)
        self.accept()


class TextoSimplesDialog(QDialog):
    def __init__(self, titulo, label, parent=None):
        super().__init__(parent)
        self.setWindowTitle(titulo)
        self.valor = None
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(label))
        self.le = QLineEdit()
        layout.addWidget(self.le)
        self.err = QLabel("")
        self.err.setStyleSheet("color:#ff5555; font-weight:bold;")
        layout.addWidget(self.err)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._ok)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _ok(self):
        v = self.le.text().strip()
        if v == "":
            self.err.setText("O valor não pode ser vazio.")
            return
        self.valor = v
        self.accept()


class PlanilhaApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Editor de Planilhas (PySide6)")
        self.resize(1100, 700)

        self.df = pd.DataFrame()
        self.sort_direcao = {}
        self.file_loaded = False

        central = QWidget()
        self.setCentralWidget(central)
        vbox = QVBoxLayout(central)
        vbox.setContentsMargins(12, 12, 12, 12)

        hb = QHBoxLayout()
        hb.setSpacing(12)

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        ICONS_DIR = os.path.join(BASE_DIR, "public", "icons")

        def make_btn(text: str, icon_file: str, slot):
            icon_path = os.path.join(ICONS_DIR, icon_file)
            btn = QPushButton(text)
            # carrega ícone se existir; caso contrário, segue só com texto
            if os.path.exists(icon_path):
                btn.setIcon(QIcon(icon_path))
                btn.setIconSize(QtCore.QSize(20, 20))
            else:
                # fallback leve para não quebrar execução
                btn.setToolTip(f"Ícone não encontrado: {icon_path}")
            btn.clicked.connect(slot)
            btn.setMinimumWidth(160)
            return btn

        hb.addWidget(make_btn("Abrir Planilha", "folder-open-solid-full.svg", self.abrir_arquivo))
        hb.addWidget(make_btn("Inserir Linha", "grip-lines-solid-full.svg", self.adicionar_linha))
        hb.addWidget(make_btn("Remover Linha", "trash-can-solid-full.svg.svg", self.remover_linha))
        hb.addWidget(make_btn("Adicionar Coluna", "table-columns-solid-full.svg", self.adicionar_coluna))
        hb.addWidget(make_btn("Remover Coluna", "trash-can-solid-full.svg", self.remover_coluna))
        hb.addWidget(make_btn("Filtros", "filter-solid-full.svg", self.editar_filtros_colunas))
        hb.addStretch()
        hb.addWidget(make_btn("Salvar", "floppy-disk-solid-full.svg", self.salvar_planilha))
        vbox.addLayout(hb)

        self.view = QTableView()
        self.view.setAlternatingRowColors(True)
        self.view.horizontalHeader().setStretchLastSection(True)
        self.view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.view.setSelectionBehavior(QTableView.SelectRows)
        self.view.setEditTriggers(QTableView.DoubleClicked | QTableView.EditKeyPressed | QTableView.AnyKeyPressed)
        self.view.horizontalHeader().sectionClicked.connect(self._ordenar_por_cabecalho)
        vbox.addWidget(self.view)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self._atualizar_status()

        self.model = PandasModel(self.df, self)
        self.view.setModel(self.model)

    def _atualizar_status(self):
        if self.df.empty:
            self.status.showMessage("Nenhuma planilha carregada")
        else:
            self.status.showMessage(f"Linhas: {len(self.df)}    Colunas: {len(self.df.columns)}")

    def _recarregar_modelo(self):
        self.model.update_df(self.df)
        self._atualizar_status()
        self.view.resizeColumnsToContents()

    # --------- Ações ----------
    def abrir_arquivo(self):
        caminho, _ = QFileDialog.getOpenFileName(self, "Selecione o arquivo Excel",
                                                 "", "Planilhas do Excel (*.xls *.xlsx)")
        if not caminho:
            return
        try:
            ext = os.path.splitext(caminho)[1].lower()
            engine = "openpyxl" if ext == ".xlsx" else "xlrd"
            self.df = pd.read_excel(caminho, engine=engine)
            self.df.attrs.setdefault("filtros", {})
            self.file_loaded = True
            self.sort_direcao.clear()
            self._recarregar_modelo()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao ler o arquivo:\n{e}")

    def _ordenar_por_cabecalho(self, logical_index: int):
        if self.df.empty or logical_index < 0 or logical_index >= len(self.df.columns):
            return
        coluna = self.df.columns[logical_index]
        direcao = self.sort_direcao.get(coluna, False)
        try:
            self.df = self.df.sort_values(by=coluna, ascending=not direcao).reset_index(drop=True)
            self.sort_direcao[coluna] = not direcao
            self._recarregar_modelo()
        except Exception as e:
            QMessageBox.warning(self, "Aviso", f"Não foi possível ordenar por '{coluna}':\n{e}")

    def adicionar_linha(self):
        if self.df.empty:
            QMessageBox.warning(self, "Aviso", "Abra uma planilha antes.")
            return

        dlg = InserirLinhaDialog(self.df.columns, self)
        if dlg.exec() != QDialog.Accepted:
            return

        pos = dlg.pos
        if pos < 0 or pos >= len(self.df):
            QMessageBox.warning(self, "Aviso", "Índice fora do intervalo.")
            return

        linha_nova = pd.DataFrame([dlg.valores], columns=self.df.columns)

        if dlg.antes:
            parte1 = self.df.iloc[:pos]
            parte2 = self.df.iloc[pos:]
        else:
            parte1 = self.df.iloc[:pos+1]
            parte2 = self.df.iloc[pos+1:]

        self.df = pd.concat([parte1, linha_nova, parte2], ignore_index=True)
        self._recarregar_modelo()

    def remover_linha(self):
        if self.df.empty:
            QMessageBox.warning(self, "Aviso", "Abra uma planilha antes.")
            return

        dlg = RemoverLinhaDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return

        idx = dlg.idx
        if idx < 0 or idx >= len(self.df):
            QMessageBox.warning(self, "Aviso", "Índice fora do intervalo.")
            return

        self.df = self.df.drop(index=idx).reset_index(drop=True)
        self._recarregar_modelo()

    def adicionar_coluna(self):
        if self.df.empty:
            QMessageBox.warning(self, "Aviso", "Abra uma planilha antes.")
            return

        dlg = TextoSimplesDialog("Adicionar Coluna", "Nome da nova coluna:", self)
        if dlg.exec() != QDialog.Accepted:
            return
        nome = dlg.valor
        if nome in self.df.columns:
            QMessageBox.warning(self, "Aviso", "Coluna já existe.")
            return
        self.df[nome] = ""
        self._recarregar_modelo()

    def remover_coluna(self):
        if self.df.empty:
            QMessageBox.warning(self, "Aviso", "Abra uma planilha antes.")
            return

        dlg = TextoSimplesDialog("Remover Coluna", "Nome da coluna a remover:", self)
        if dlg.exec() != QDialog.Accepted:
            return
        nome = dlg.valor
        if nome not in self.df.columns:
            QMessageBox.warning(self, "Aviso", "Coluna não existe.")
            return
        self.df = self.df.drop(columns=[nome])
        self._recarregar_modelo()

    def editar_filtros_colunas(self):
        if self.df.empty:
            QMessageBox.warning(self, "Aviso", "Abra uma planilha antes.")
            return

        filtros_atual = self.df.attrs.get("filtros", {})
        dlg = FiltrosDialog(self.df.columns, filtros_atual, self)
        novos = dlg.get_filters()
        if novos is None:
            return

        try:
            df_filtrado = self.df.copy()
            for col, val in novos.items():
                df_filtrado = df_filtrado[df_filtrado[col].astype(str).str.contains(val, case=False, na=False)]
            self.df.attrs["filtros"] = novos
            self.df = df_filtrado.reset_index(drop=True)
            self._recarregar_modelo()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao aplicar filtros:\n{e}")

    def salvar_planilha(self):
        if self.df.empty:
            QMessageBox.warning(self, "Aviso", "Nenhuma planilha para salvar.")
            return

        caminho, _ = QFileDialog.getSaveFileName(self, "Salvar planilha", "", "Planilhas Excel (*.xlsx)")
        if not caminho:
            return
        if not caminho.lower().endswith(".xlsx"):
            caminho += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Planilha"

            # Cabeçalhos em negrito
            for idx, col in enumerate(self.df.columns, 1):
                cell = ws.cell(row=1, column=idx, value=col)
                cell.font = Font(bold=True)

            # Dados
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Tabela estilizada
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tab = Table(displayName="Tabela1", ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            wb.save(caminho)
            QMessageBox.information(self, "Sucesso", "Planilha salva com sucesso.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar arquivo:\n{e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)

    # 🌙 Dark Mode global
    app.setStyleSheet("""
QMainWindow { background-color: #2b2b2b; }
QPushButton {
    background-color: #007acc;
    color: white;
    border-radius: 6px;
    padding: 6px 12px;
    font-weight: bold;
}
QPushButton:hover { background-color: #3399ff; }
QTableView {
    background-color: #1e1e1e;
    color: #eeeeee;
    gridline-color: #444444;
    selection-background-color: #3399ff;
    selection-color: white;
    alternate-background-color: #2e2e2e;
}
QHeaderView::section {
    background-color: #3c3c3c;
    color: white;
    padding: 4px;
    border: 1px solid #444;
}
QLineEdit {
    background-color: #3c3c3c;
    color: white;
    border: 1px solid #555;
    border-radius: 4px;
    padding: 4px;
}
QLabel { color: #eeeeee; }
QStatusBar { background-color: #333333; color: #bbbbbb; }
QDialog { background-color: #2b2b2b; }
""")

    win = PlanilhaApp()
    win.show()
    sys.exit(app.exec())
